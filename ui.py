# ui.py

import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import ImageTk
from openpyxl import load_workbook
from card_builder import generar_tarjeta, generar_imagen_previa


class TarjetaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador de Tarjetas - Zebra ZXP3")

        # Variables
        self.plantilla_var = tk.StringVar()
        self.foto_var = tk.StringVar()
        self.nombre_var = tk.StringVar()
        self.legajo_var = tk.StringVar()
        self.dni_var = tk.StringVar()

        self.preview_image = None
        self.carpeta_fotos = ""
        self.rows = []

        self._build_ui()

    def _build_ui(self):
        self.root.columnconfigure(0, weight=0)  # columna izquierda (formulario)
        self.root.columnconfigure(1, weight=1)  # columna derecha (tabla + preview)
        self.root.rowconfigure(0, weight=1)

        # ─────────────────────────────────────────
        # FRAME IZQUIERDO: formulario + botones
        # ─────────────────────────────────────────
        left_frame = tk.Frame(self.root)
        left_frame.grid(row=0, column=0, sticky="nsw", padx=10, pady=10)

        padding = {"padx": 5, "pady": 3}

        # Plantilla
        tk.Label(left_frame, text="Plantilla de tarjeta:").grid(row=0, column=0, sticky="w", **padding)
        tk.Entry(left_frame, textvariable=self.plantilla_var, width=40).grid(row=1, column=0, **padding)
        tk.Button(left_frame, text="Buscar plantilla...", command=self.seleccionar_plantilla).grid(
            row=1, column=1, **padding
        )

        # Foto
        tk.Label(left_frame, text="Foto individual:").grid(row=2, column=0, sticky="w", **padding)
        tk.Entry(left_frame, textvariable=self.foto_var, width=40).grid(row=3, column=0, **padding)
        tk.Button(left_frame, text="Buscar foto...", command=self.seleccionar_foto).grid(
            row=3, column=1, **padding
        )

        # Datos
        tk.Label(left_frame, text="Nombre y Apellido:").grid(row=4, column=0, sticky="w", **padding)
        tk.Entry(left_frame, textvariable=self.nombre_var, width=40).grid(row=5, column=0, columnspan=2, **padding)

        tk.Label(left_frame, text="Legajo:").grid(row=6, column=0, sticky="w", **padding)
        tk.Entry(left_frame, textvariable=self.legajo_var, width=20).grid(row=7, column=0, sticky="w", **padding)

        tk.Label(left_frame, text="DNI:").grid(row=8, column=0, sticky="w", **padding)
        tk.Entry(left_frame, textvariable=self.dni_var, width=20).grid(row=9, column=0, sticky="w", **padding)

        # Botones individuales
        tk.Button(left_frame, text="Vista previa individual", command=self.actualizar_previa).grid(
            row=10, column=0, columnspan=2, pady=8
        )

        tk.Button(left_frame, text="Generar PDF individual", command=self.generar_pdf, bg="#4CAF50", fg="white").grid(
            row=11, column=0, columnspan=2, pady=8
        )

        # Botones masivos
        tk.Label(left_frame, text="Exportación masiva desde Excel").grid(
            row=12, column=0, columnspan=2, sticky="w", pady=(15, 3)
        )

        tk.Button(left_frame, text="Cargar Excel + fotos", command=self.cargar_excel).grid(
            row=13, column=0, columnspan=2, pady=5
        )

        tk.Button(left_frame, text="Exportar TODOS los PDFs del Excel", command=self.exportar_todos).grid(
            row=14, column=0, columnspan=2, pady=5
        )

        # ─────────────────────────────────────────
        # FRAME DERECHO: tabla + vista previa
        # ─────────────────────────────────────────
        right_frame = tk.Frame(self.root)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        right_frame.rowconfigure(0, weight=1)
        right_frame.rowconfigure(1, weight=0)
        right_frame.columnconfigure(0, weight=1)

        # Tabla de datos del Excel
        self.tree = ttk.Treeview(right_frame, columns=("nombre", "legajo", "dni", "foto"), show="headings")
        self.tree.heading("nombre", text="Nombre")
        self.tree.heading("legajo", text="Legajo")
        self.tree.heading("dni", text="DNI")
        self.tree.heading("foto", text="Foto")

        self.tree.column("nombre", width=180)
        self.tree.column("legajo", width=80)
        self.tree.column("dni", width=100)
        self.tree.column("foto", width=180)

        self.tree.grid(row=0, column=0, sticky="nsew")

        # Scrollbar para la tabla
        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Evento: al seleccionar una fila → vista previa (misma zona que la individual)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # Vista previa (única, compartida)
        self.preview_label = tk.Label(right_frame, text="Vista previa")
        self.preview_label.grid(row=1, column=0, columnspan=2, pady=10)

    # ─────────────────────────────────────────
    # Funciones de selección de archivos
    # ─────────────────────────────────────────
    def seleccionar_plantilla(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar plantilla",
            filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg;*.bmp")]
        )
        if ruta:
            self.plantilla_var.set(ruta)

    def seleccionar_foto(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar foto",
            filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg;*.bmp")]
        )
        if ruta:
            self.foto_var.set(ruta)

    # ─────────────────────────────────────────
    # Vista previa individual
    # ─────────────────────────────────────────
    def actualizar_previa(self):
        tarjeta = generar_imagen_previa(
            self.plantilla_var.get(),
            self.foto_var.get(),
            self.nombre_var.get(),
            self.legajo_var.get(),
            self.dni_var.get()
        )

        preview = tarjeta.resize((250, 400))
        self.preview_image = ImageTk.PhotoImage(preview)
        self.preview_label.config(image=self.preview_image)

    # ─────────────────────────────────────────
    # Generación PDF individual
    # ─────────────────────────────────────────
    def generar_pdf(self):
        try:
            salida_pdf = generar_tarjeta(
                self.plantilla_var.get(),
                self.foto_var.get(),
                self.nombre_var.get(),
                self.legajo_var.get(),
                self.dni_var.get()
            )
            messagebox.showinfo("Éxito", f"PDF generado:\n{salida_pdf}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error:\n{e}")

    # ─────────────────────────────────────────
    # Carga de Excel y listado
    # ─────────────────────────────────────────
    def cargar_excel(self):
        ruta_excel = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not ruta_excel:
            return

        carpeta_fotos = filedialog.askdirectory(
            title="Seleccionar carpeta de fotos"
        )
        if not carpeta_fotos:
            return

        self.carpeta_fotos = carpeta_fotos

        # Leer Excel
        wb = load_workbook(ruta_excel)
        ws = wb.active

        self.rows = []
        self.tree.delete(*self.tree.get_children())

        for row in ws.iter_rows(min_row=2, values_only=True):
            self.rows.append(row)
            self.tree.insert("", "end", values=row)

        messagebox.showinfo("OK", "Excel cargado. Seleccioná una fila para ver la vista previa.")

    # ─────────────────────────────────────────
    # Vista previa desde selección en la tabla
    # ─────────────────────────────────────────
    def on_tree_select(self, event):
        selected = self.tree.selection()
        if not selected:
            return

        item = self.tree.item(selected[0])
        nombre, legajo, dni, foto = item["values"]

        # Buscar foto
        foto_path = ""
        if foto:
            posible = os.path.join(self.carpeta_fotos, foto)
            if os.path.exists(posible):
                foto_path = posible

        tarjeta = generar_imagen_previa(
            self.plantilla_var.get(),
            foto_path,
            nombre or "",
            str(legajo) or "",
            str(dni) or ""
        )

        preview = tarjeta.resize((250, 400))
        self.preview_image = ImageTk.PhotoImage(preview)
        self.preview_label.config(image=self.preview_image)

    # ─────────────────────────────────────────
    # Exportación masiva
    # ─────────────────────────────────────────
    def exportar_todos(self):
        if not self.rows:
            messagebox.showerror("Error", "Primero cargá un Excel.")
            return

        for nombre, legajo, dni, foto in self.rows:
            foto_path = ""
            if foto:
                posible = os.path.join(self.carpeta_fotos, foto)
                if os.path.exists(posible):
                    foto_path = posible

            generar_tarjeta(
                self.plantilla_var.get(),
                foto_path,
                nombre or "",
                str(legajo) or "",
                str(dni) or ""
            )

        messagebox.showinfo("Éxito", "Todas las tarjetas fueron exportadas.")
