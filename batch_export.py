# batch_export.py

import os
from openpyxl import load_workbook
from card_builder import generar_tarjeta

def exportar_masivo(ruta_excel, carpeta_fotos, plantilla_path=""):
    """
    ruta_excel: archivo .xlsx con columnas A-D
    carpeta_fotos: carpeta donde están las fotos
    plantilla_path: opcional, plantilla para todas las tarjetas
    """

    wb = load_workbook(ruta_excel)
    ws = wb.active

    resultados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre, legajo, dni, archivo_foto = row

        # Construir ruta de foto
        foto_path = ""
        if archivo_foto:
            posible = os.path.join(carpeta_fotos, archivo_foto)
            if os.path.exists(posible):
                foto_path = posible

        # Generar PDF
        salida = generar_tarjeta(
            plantilla_path,
            foto_path,
            nombre if nombre else "",
            str(legajo) if legajo else "",
            str(dni) if dni else ""
        )

        resultados.append(salida)

    return resultados
